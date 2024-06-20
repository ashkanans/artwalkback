import json
import os
from datetime import datetime

from openpyxl.styles import Border, Side, DEFAULT_FONT
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from backend.scraper.codal.service.cells_service import CellsService
from backend.scraper.fipiran.service.index_service import IndexService
from backend.scraper.logger.base_logger import BaseLogger
from backend.scraper.tsetmc.service.closing_price_daily_list_service import ClosingPriceDailyListService
from backend.scraper.tsetmc.service.instrument_info_service import InstrumentInfoService
from backend.utils.export.excel.utils import excel_cell_to_row_col, calculate_proper_address, \
    filtered_cells_by_monthly_performance, make_cells_custumized_based_on_headers, \
    make_sector_cells_custumized_based_on_headers, make_IRRTD_custumized_based_on_headers


class CustomWorkbook(BaseLogger):
    def __init__(self, config_path):

        super().__init__()
        script_dir = os.path.dirname(os.path.realpath(__file__))
        config_path = os.path.join(script_dir, config_path)

        with open(config_path, 'r', encoding='utf-8') as config_file:
            self.config = json.load(config_file)

        self.wb = Workbook()

    def add_sheets(self):
        for sheet in self.config.get("sheets"):
            name = sheet.get('nameFa')
            symbol = sheet.get('symbol')

            self.wb.create_sheet(symbol + " - " + name)

    def create_excel(self):
        sheet_name_to_remove = 'Sheet'

        # Check if the sheet exists before attempting to remove
        if sheet_name_to_remove in self.wb.sheetnames:
            # Get the sheet
            sheet_to_remove = self.wb[sheet_name_to_remove]

            # Remove the sheet
            self.wb.remove(sheet_to_remove)

            # Save the workbook
            current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Add the date and time to the output file name
            output_file_name = f"{self.config.get('output-file-name')}_{current_datetime}.xlsx"

            DEFAULT_FONT.name = "Arial"
            # Save the workbook with the updated file name
            self.wb.save(output_file_name)
            print(f"The sheet '{sheet_name_to_remove}' has been removed.")
        else:
            print(f"The sheet '{sheet_name_to_remove}' does not exist in the workbook.")

    def set_sheets_properties(self):
        sheet_props = self.config.get("sheet-properties")
        for key, value in sheet_props.items():
            sheetNames = self.wb.sheetnames
            for sheetName in sheetNames:
                sheet = self.wb.get_sheet_by_name(sheetName)
                if key == "right-to-left":
                    sheet.sheet_view.rightToLeft = value

    def fill_sheets(self):
        for sheet in self.config.get("sheets"):

            name = sheet.get('nameFa')
            symbol = sheet.get('symbol')

            worksheet = self.wb.get_sheet_by_name(symbol + " - " + name)

            if sheet.get("nameEn") == "Last Codal Announcement":
                cellsService = CellsService()
                letters = cellsService.get_cells_by_id_and_symbol(symbol)

                if len(letters) == 0:
                    self.logger.error(f"No letters found for symbol: {symbol}")
                    break

                final_letter_tracingNo, final_letter_cell_list = next(iter(letters.items()))

                # Define cell border style
                border_style = Border(left=Side(style='thin'),
                                      right=Side(style='thin'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))
                # Define fonts for sheet
                regualr_Font = Font(name='Tahoma', size=10)
                bold_Font = Font(name='Tahoma', size=9.0, bold=True)

                headerRowIndexes = [1, 6, 39, 41, 42, 43]

                for cells in final_letter_cell_list:
                    last_written_row = worksheet.max_row if worksheet.max_row else 1
                    for cell in cells:
                        column, row = excel_cell_to_row_col(cell.address)
                        # Construct the new address by appending the current row
                        new_address = calculate_proper_address(column, row, last_written_row, "row")

                        # Check first column or row 1 ro row 2 to change font
                        if column == 'A' or row in [1, 2]:
                            worksheet[new_address].font = bold_Font  # Set the font to bold
                        else:
                            worksheet[new_address].font = regualr_Font
                        try:

                            if column == 'J' and row == 4:
                                a = 0
                            # Try to convert the cell value to a number
                            if cell.value:
                                if cell.value.isdigit():
                                    converted_value = float(cell.value)
                                    if converted_value == 0:
                                        worksheet[new_address].number_format = '0'
                                    else:
                                        worksheet[new_address].number_format = '0,000'

                                else:
                                    if '-' in cell.value:
                                        converted_value = cell.value.replace('-', '')
                                        if converted_value.isdigit():
                                            converted_value = float(converted_value)
                                    else:
                                        converted_value = cell.value

                                worksheet[new_address] = converted_value


                            else:
                                worksheet[new_address] = ''

                            if column == 'A':
                                if headerRowIndexes.__contains__(row):
                                    worksheet[new_address].alignment = worksheet[new_address].alignment.copy(
                                        horizontal='center', vertical='center')
                                else:
                                    worksheet[new_address].alignment = worksheet[new_address].alignment.copy(
                                        horizontal='right', vertical='center')
                            else:
                                worksheet[new_address].alignment = worksheet[new_address].alignment.copy(
                                    horizontal='center', vertical='center')
                            worksheet[new_address].border = border_style

                            # Adjust the width of the column
                            max_length = max(worksheet.column_dimensions[column].width,
                                             len(str(worksheet[new_address].value)))
                            worksheet.column_dimensions[column].width = max_length


                        except ValueError:
                            # If an exception occurs (e.g., if the cell value is not a number),
                            # simply use the original cell value
                            # worksheet[new_address] = cell.value
                            print(ValueError)
                # Merge Cells
                worksheet.merge_cells('C1:F1')
                worksheet.merge_cells('G1:I1')
                worksheet.merge_cells('J1:M1')
                worksheet.merge_cells('N1:Q1')
                worksheet.merge_cells('R1:U1')
                worksheet.merge_cells('V1:Y1')


            elif sheet.get("nameEn") == "Monthly Performance":
                monthly_performance_headers = sheet.get('monthly-performance-headers')
                cellsService = CellsService()
                letters = cellsService.get_cells_by_id_and_symbol(symbol)

                if len(letters) == 0:
                    self.logger.error(f"No letters found for symbol: {symbol}")
                    break

                i = 0
                for tracingNo, letter in letters.items():
                    first_table_cells = letter[0]
                    filetered_cells = filtered_cells_by_monthly_performance(first_table_cells)

                    custumzied_filtered_cells = make_cells_custumized_based_on_headers(filetered_cells,
                                                                                       monthly_performance_headers)

                    if i > 0:
                        custumzied_filtered_cells = [cell for cell in custumzied_filtered_cells if
                                                     'A' not in cell.address]

                    last_written_column = worksheet.max_column if worksheet.max_column else 1

                    for cell in custumzied_filtered_cells:
                        column, row = excel_cell_to_row_col(cell.address)
                        # Construct the new address by appending the current column
                        new_address = calculate_proper_address(column, row, last_written_column, "column")

                        try:
                            # Try to convert the cell value to a number
                            if cell.value:
                                double_value = float(cell.value)
                                if double_value == 0.0:
                                    worksheet[new_address] = ""
                                else:
                                    worksheet[new_address] = double_value
                                    worksheet[new_address].number_format = '0,000'
                                    worksheet[new_address].alignment = worksheet[new_address].alignment.copy(
                                        horizontal='center', vertical='center')
                        except ValueError:
                            # If an exception occurs (e.g., if the cell value is not a number),
                            # simply use the original cell value
                            worksheet[new_address] = cell.value

                    i += 1


            elif sheet.get("nameEn") == "Sector Prices History":
                sector_prices_headers = sheet.get('sector-prices-headers')
                instrumentService = InstrumentInfoService()
                table_headers = sheet.get('sector-prices-and_indexs')

                indexes = sheet.get("sector-indexes")
                instruments = []
                for index in indexes:
                    instruments.append(instrumentService.get_instrument_info_by_persian_symbol(index))

                closingPriceDailyTables = []
                for instrument in instruments:
                    closingPriceDailyTables.append(f"tse_inst_{instrument.insCode}_history")

                listOfClosingPrice = []
                for table in closingPriceDailyTables:
                    closingPriceDailyListService = ClosingPriceDailyListService(table)
                    listOfClosingPrice.append(closingPriceDailyListService.get_closing_price_daily_list_last_30_items())

                indexService = IndexService()
                fipiran_index = indexService.get_all_indices()
                custumzied_filtered_cells = make_sector_cells_custumized_based_on_headers(indexes,
                                                                                          listOfClosingPrice,
                                                                                          sector_prices_headers,
                                                                                          fipiran_index, table_headers)

                for cell in custumzied_filtered_cells:
                    new_address = cell.address
                    try:
                        # Try to convert the cell value to a number
                        if cell.value:
                            double_value = float(cell.value)
                            if double_value == 0.0:
                                worksheet[new_address] = ""
                            else:
                                worksheet[new_address] = double_value
                                worksheet[new_address].number_format = '0,000'
                                worksheet[new_address].alignment = worksheet[new_address].alignment.copy(
                                    horizontal='center', vertical='center')
                    except ValueError:
                        # If an exception occurs (e.g., if the cell value is not a number),
                        # simply use the original cell value
                        worksheet[new_address] = cell.value


            elif sheet.get("nameEn") == "Industry-related Real-time Data":
                indexes = sheet.get('Industry-related-indexes')
                headers = sheet.get('headers')

                custumzied_filtered_cells = make_IRRTD_custumized_based_on_headers(indexes,
                                                                                   headers)

                for cell in custumzied_filtered_cells:
                    new_address = cell.address
                    try:
                        # Try to convert the cell value to a number
                        if cell.value:
                            double_value = float(cell.value)
                            if double_value == 0.0:
                                worksheet[new_address] = ""
                            else:
                                worksheet[new_address] = double_value
                                worksheet[new_address].number_format = '0,000'
                                worksheet[new_address].alignment = worksheet[new_address].alignment.copy(
                                    horizontal='center', vertical='center')
                    except ValueError:
                        # If an exception occurs (e.g., if the cell value is not a number),
                        # simply use the original cell value
                        worksheet[new_address] = cell.value


            elif sheet.get("nameEn") == "Profit and Loss Statement for the Last 5 Years":
                pass
