from openpyxl.workbook import Workbook

from backend.scraper.codal.model.letter_headers import convert_to_format
from backend.scraper.codal.service.inventory_letter_headers_service import InventoryLetterHeadersService
from backend.scraper.codal.service.letter_headers_service import LettersRowHeadersService
from backend.scraper.logger.base_logger import BaseLogger
from backend.scraper.tsetmc.service.closing_price_daily_list_service import ClosingPriceDailyListService
from backend.scraper.tsetmc.service.instrument_group_service import InstrumentGroupService
from backend.scraper.tsetmc.service.instrument_info_service import InstrumentInfoService
from backend.utils.export.excel.utils import *
from backend.utils.scraper.realtimedata.utils import convert_persian_to_arabic


def get_header_cells(cells):
    cell_domestic_sales = None
    cell_export_sales = None
    cell_service_income = None
    cell_sales_return = None

    for cell in cells:
        if 'فروش داخلی' in cell.value and 'جمع' not in cell.value:
            cell_domestic_sales = cell
        elif 'فروش صادراتی' in cell.value and 'جمع' not in cell.value:
            cell_export_sales = cell
        elif 'درآمد ارائه خدمات' in cell.value and 'جمع' not in cell.value:
            cell_service_income = cell
        elif 'برگشت از فروش' in cell.value and 'جمع' not in cell.value:
            cell_sales_return = cell

    return [cell_domestic_sales, cell_export_sales, cell_service_income, cell_sales_return]


class DummyWorkbook(BaseLogger):
    def __init__(self):
        super().__init__()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def fill_dummy_sheet_with_final_mpr_with_cells(self, final_letter_cell_list):
        for cells in final_letter_cell_list:
            last_written_row = self.worksheet.max_row if self.worksheet.max_row else 1
            for cell in cells:
                column, row = excel_cell_to_row_col(cell.address)
                # Construct the new address by appending the current row
                new_address = calculate_proper_address(column, row, last_written_row, "row")

                # Check first column or row 1 ro row 2 to change font

                try:

                    # Try to convert the cell value to a number
                    if cell.value:
                        if cell.value.isdigit():
                            converted_value = float(cell.value)
                            if converted_value == 0:
                                self.worksheet[new_address].number_format = '0'
                            else:
                                self.worksheet[new_address].number_format = '0,000'

                        else:
                            if '-' in cell.value:
                                converted_value = cell.value.replace('-', '')
                                if converted_value.isdigit():
                                    converted_value = float(converted_value)
                            else:
                                converted_value = cell.value

                        self.worksheet[new_address] = converted_value


                    else:
                        self.worksheet[new_address] = ''




                except ValueError:
                    # If an exception occurs (e.g., if the cell value is not a number),
                    # simply use the original cell value
                    # worksheet[new_address] = cell.value
                    print(ValueError)

    def create_excel(self):

        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")

        output_file_name = f"test_{current_datetime}.xlsx"

        self.workbook.save(output_file_name)
        print(f"The workbook '{output_file_name}' has been created.")

    def get_rows_from_workbook(self):
        rows_list = []

        for row in self.worksheet.iter_rows(values_only=True):
            rows_list.append(row)

        return rows_list

    def fill_with_cells(self, cells):
        for cell in cells:
            self.worksheet[cell.address] = cell.value

    def fill_dummy_sheet_with_monthly_performance_with_cells(self, letters, symbol):
        lettersRowHeadersService = LettersRowHeadersService()
        headers = lettersRowHeadersService.get_entry_by_symbol(symbol)
        headers_dict = convert_to_format(headers)

        i = 0
        for tracingNo, letter in letters.items():

            first_table_cells = letter[0]
            filetered_cells = filtered_cells_by_monthly_performance(first_table_cells)

            header_cells = get_header_cells(filetered_cells)
            if any(cell is None for cell in header_cells):
                continue
            custumzied_filtered_cells = make_cells_custumized_based_on_headers(filetered_cells,
                                                                               headers_dict,
                                                                               header_cells)

            if i > 0:
                custumzied_filtered_cells = [cell for cell in custumzied_filtered_cells if
                                             'A' not in cell.address]

            last_written_column = self.worksheet.max_column if self.worksheet.max_column else 1

            for cell in custumzied_filtered_cells:
                column, row = excel_cell_to_row_col(cell.address)
                # Construct the new address by appending the current column
                new_address = calculate_proper_address(column, row, last_written_column, "column")

                self.worksheet[new_address] = cell.value

            i += 1

    def fill_dummy_sheet_with_sector_prices_with_cells(self, sector_code, indexes):
        instrumentService = InstrumentInfoService()
        instrumentGroupService = InstrumentGroupService()

        sector_prices_headers = {
            "datetime_str": "تاریخ",
            "priceChange": "اولین",
            "pDrCotVal": "آخرین معامله",
            "pClosing": "کمترین",
            "priceFirst": "بیشترین",
            "qTotTran5J": "حجم"
        }
        table_headers = [
            "تاریخ",
            "شاخص کل",
        ]
        group = instrumentGroupService.get_instrument_group_by_group_code(sector_code)

        instrumentService = InstrumentInfoService()
        instruments = []
        for index in indexes:
            instruments.append(
                instrumentService.get_instrument_info_by_persian_symbol(index.replace("ک", "ك").replace("ی", "ي")))

        closingPriceDailyTables = []
        for instrument in instruments:
            closingPriceDailyTables.append(f"tse_inst_{instrument.insCode}_history")

        listOfClosingPrice = []
        for table in closingPriceDailyTables:
            closingPriceDailyListService = ClosingPriceDailyListService(table)
            listOfClosingPrice.append(closingPriceDailyListService.get_closing_price_daily_list_last_30_items())

        indexService = IndexService()
        fipiran_index = {}
        for header in table_headers:
            header_org = header
            if 'شاخص' in header:
                header = header.replace("شاخص", "").strip()
                fipiran_index[header_org] = indexService.get_by_nameFa(convert_persian_to_arabic(header))
        fipiran_index["شاخص" + " " + group.GroupName] = indexService.get_index_by_sectorCode(sector_code)
        table_headers.append("شاخص" + " " + group.GroupName)

        custumzied_filtered_cells = make_sector_cells_custumized_based_on_headers(indexes,
                                                                                  listOfClosingPrice,
                                                                                  sector_prices_headers,
                                                                                  fipiran_index, table_headers)
        for cell in custumzied_filtered_cells:
            new_address = cell.address
            self.worksheet[new_address] = cell.value

    def fill_dummy_sheet_with_income_statement_with_cells(self, letters, symbol):
        count = 0
        main_header = []

        for tracingNo, letter in letters.items():

            first_table_cells = letter[0]
            letter_header = get_first_Column_of_table(first_table_cells)
            difference = [value for dictionary in letter_header for value in dictionary.values() if
                          value not in main_header]
            main_header = main_header + difference
            if count == 0:
                for i, string in enumerate(main_header, start=3):
                    self.worksheet[f'A{i}'] = string
            else:
                # Insert new rows at the end of the worksheet
                self.worksheet.insert_rows(idx=self.worksheet.max_row + 1, amount=len(difference))

                # Add values from your list to the new rows
                for i, string in enumerate(difference, start=self.worksheet.max_row + 1):
                    self.worksheet[f'A{i}'] = string

            max_column = self.worksheet.max_column
            for cell in first_table_cells:
                if 'A' not in cell.address:
                    header_title = self.get_value(letter_header, cell.rowSequence)
                    if header_title:
                        self.fill_sheet_by_header_name(cell, self.get_new_cell_address(cell.address, max_column),header_title)
                    else:
                        self.worksheet[self.get_new_cell_address(cell.address, max_column)] = cell.value

            count += 1

    def find_in_worksheet(self, column, value):
        for row in range(1, self.worksheet.max_row + 1):
            if self.worksheet[column + str(row)].value == value:
                return row
        return None

    def fill_sheet_by_header_name(self, cell,column, header):
        row = self.find_in_worksheet('A', header)
        if row:
            word, number = self.split_letter_and_number(column)
            new_cell_address = word + str(row)
            self.worksheet[new_cell_address] = cell.value

    def get_value(self, list_of_dicts, key):
        for dictionary in list_of_dicts:
            if key in dictionary:
                return dictionary[key]
        return None

    def get_new_cell_address(self, cell_adderss, max_column):
        word_num_map = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8, 'I': 9, 'J': 10,
                        'K': 11, 'L': 12, 'M': 13, 'N': 14, 'O': 15, 'P': 16, 'Q': 17, 'R': 18, 'S': 19,
                        'T': 20, 'U': 21, 'V': 22, 'W': 23, 'X': 24, 'Y': 25, 'Z': 26}

        result = self.split_letter_and_number(cell_adderss)
        word, row_num = result
        i = max_column + word_num_map[word] - 2

        if i < 26:
            column_name = chr(65 + (i))
        else:
            first_letter = chr(65 + ((i) // 26 - 1))
            second_letter = chr(65 + ((i) % 26))
            column_name = f"{first_letter}{second_letter}"
        column_address = column_name + row_num

        return column_address

    def split_letter_and_number(self, s):
        match = re.match(r"([a-zA-Z]+)(\d+)", s)
        if match:
            return match.groups()
        else:
            return None

    def get_invertory_table_headers(self, letters, letterType):

        headers = []
        headers_type = []
        for tracingNo, tables in letters.items():
            counter = 0
            for table in tables:
                for cell in table:
                    if 'A' in cell.address:
                        if cell.value:
                            if cell.value not in headers and 'شرح' not in cell.value and 'اوليه' not in cell.value and 'جمع' not in cell.value:
                                headers.append(cell.value)
                                headers_type.append(letterType[counter])

                                result = self.split_letter_and_number(cell.address)
                                word, row_num = result

                counter += 1

        for i in range(0, len(headers)):
            self.worksheet["A" + str(i + 1)] = headers[i]
            self.worksheet["B" + str(i + 1)] = headers_type[i]

    def fill_dummy_sheet_with_inventory_letter_with_cells(self, letters, symbol):

        counter = 0
        product_headers = []
        time_header = ''
        time_header_address = 'D1'
        inventory_letter_header_service = InventoryLetterHeadersService()
        main_product_headers = inventory_letter_header_service.get_all_symbol_headers_by_type(symbol,
                                                                                              'گردش مقداری - ریالی موجودی کالا')
        main_requirment_headers = inventory_letter_header_service.get_all_symbol_headers_by_type(symbol,
                                                                                                 'خرید و مصرف مواد اولیه')
        last_row = '2'
        for i in [1, 2]:
            if i == 1:
                if main_product_headers:
                    self.worksheet, last_row = create_sheet_template_product(self.worksheet, main_product_headers,
                                                                             last_row)
            else:
                if main_requirment_headers:
                    last_row = str(int(last_row) + 3)
                    self.worksheet, last_row = create_sheet_template_requirement(self.worksheet,
                                                                                 main_requirment_headers, last_row)

        for tracingNo, tables in letters.items():
            counter = 0
            last_row = '2'
            for table in tables:
                # table = tables[0]
                total_dict = {}

                product_headers = get_first_Column_of_table(table)

                for cell in table:
                    if 'منتهي به' in cell.value:
                        time_header = cell.value
                    if 'A' not in cell.address and 'B' not in cell.address:
                        if cell.rowSequence not in ['1', '2', '3', '4']:
                            result = self.split_letter_and_number(cell.address)
                            word, row_num = result
                            if counter == 0:
                                dict = creat_customized_dict_product(cell.value, word)
                            else:
                                dict = creat_customized_dict_requirement(cell.value, word)
                            header = get_header_by_row_code(row_num, product_headers)
                            if header:
                                total_dict = fill_customized_dict(total_dict, header, dict)
                            # self.worksheet[cell.address] = cell.value

                self.worksheet[time_header_address] = time_header
                text_address, number_address = self.split_letter_and_number(time_header_address)

                if counter == 0:
                    self.worksheet, last_row = fill_sheet_by_customized_dict_product(self.worksheet, total_dict,
                                                                                     last_row,
                                                                                     text_address)
                else:
                    self.worksheet, last_row = fill_sheet_by_customized_dict_requirement(self.worksheet, total_dict,
                                                                                         last_row,
                                                                                         text_address)
                counter += 1
                last_row = str(int(last_row) + 3)
            time_header_address = self.get_new_cell_address('B1', self.worksheet.max_column)
